"""
NCD Tracker - Data Store
Supports two backends:
  DataStore       → local Excel (Bond_Primary_Deals.xlsx)
  GSheetDataStore → Google Sheets via service_account.json
  get_data_store()→ factory; picks backend from st.session_state
"""

import pandas as pd
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date
from pathlib import Path

import config
from constants import (
    PIPELINE_HEADERS, CLOSED_HEADERS, COLUMN_WIDTHS, NUMBER_FORMATS,
    PIPELINE_COL_ISSUER, PIPELINE_COL_DATE, PIPELINE_COL_TYPE,
    PIPELINE_COL_ISSUER_TYPE, PIPELINE_COL_QUANTUM, PIPELINE_COL_CREDIT,
    PIPELINE_COL_RATING, PIPELINE_COL_SECURITY, PIPELINE_COL_CHECKLIST,
    PIPELINE_COL_CREATED, PIPELINE_COL_STATUS,
    CLOSED_COL_ISSUER, CLOSED_COL_ISSUE_DATE, CLOSED_COL_ISIN,
    CLOSED_COL_COUPON, CLOSED_COL_TENOR, CLOSED_COL_MATURITY,
    CLOSED_COL_QUANTUM, CLOSED_COL_RATING, CLOSED_COL_TYPE,
    CLOSED_COL_ISSUER_TYPE, CLOSED_COL_SECURITY,
    parse_excel_date, parse_coupon, normalise_instrument_type,
    get_checklist_for_instrument,
)
from models import PipelineDeal, ClosedDeal, PhaseChecklist, ChecklistItem


class DataStore:
    """Manages all Excel read/write operations against Bond_Primary_Deals.xlsx."""

    def __init__(self, file_path=None):
        self.file_path = Path(file_path) if file_path else config.DATA_FILE
        self._ensure_file_exists()

    # ══════════════════════════════════════════════
    #  FILE BOOTSTRAP
    # ══════════════════════════════════════════════

    def _ensure_file_exists(self):
        if not self.file_path.exists():
            self._create_new_file()

    def _create_new_file(self):
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        self._setup_sheet_headers(wb.create_sheet(config.SHEET_PIPELINE), PIPELINE_HEADERS)
        self._setup_sheet_headers(wb.create_sheet(config.SHEET_CLOSED),   CLOSED_HEADERS)
        wb.save(self.file_path)

    def _setup_sheet_headers(self, sheet, headers):
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font      = Font(bold=True, size=11, color="FFFFFF")
            cell.fill      = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.column_dimensions[cell.column_letter].width = COLUMN_WIDTHS.get(header, 15)
        sheet.freeze_panes = 'A2'

    # ══════════════════════════════════════════════
    #  READ PIPELINE
    # ══════════════════════════════════════════════

    def load_pipeline_deals(self):
        try:
            df = pd.read_excel(self.file_path, sheet_name=config.SHEET_PIPELINE)
        except Exception:
            return []

        if df.empty:
            return []

        deals = []
        for _, row in df.iterrows():
            try:
                deal = self._row_to_pipeline(row)
                if deal:
                    deals.append(deal)
            except Exception as e:
                print(f"[DataStore] skip pipeline row: {e}")
        return deals

    def _row_to_pipeline(self, row):
        def get(col, default=None):
            val = row.get(col)
            if val is None or (hasattr(val, '__class__') and str(type(val)) == "<class 'float'>"):
                import math
                if isinstance(val, float) and math.isnan(val):
                    return default
            return val if val is not None else default

        company_name = str(get(PIPELINE_COL_ISSUER, "")).strip()
        if not company_name or company_name.lower() in ("nan", "none", ""):
            return None

        raw_type      = str(get(PIPELINE_COL_TYPE, "Listed"))
        instrument    = normalise_instrument_type(raw_type)
        issuer_type   = str(get(PIPELINE_COL_ISSUER_TYPE, "FS")).strip()
        # Derive asset_class from issuer_type as a sensible default
        asset_class   = "NBFC" if issuer_type == "FS" else "Corporate"

        raw_quantum   = get(PIPELINE_COL_QUANTUM, 0)
        try:
            issuance_size = float(str(raw_quantum).replace(',', ''))
        except Exception:
            issuance_size = 0.0

        funding_date  = parse_excel_date(get(PIPELINE_COL_DATE))
        rating        = str(get(PIPELINE_COL_RATING, "N/A") or "N/A").strip()
        security      = str(get(PIPELINE_COL_SECURITY, "Unsecured") or "Unsecured").strip()
        status        = str(get(PIPELINE_COL_STATUS, get(PIPELINE_COL_CREDIT, "In Progress")) or "In Progress").strip()
        created_raw   = get(PIPELINE_COL_CREATED)
        created_date  = parse_excel_date(created_raw) if created_raw else date.today()

        # Checklist JSON
        checklists = {}
        cl_raw = get(PIPELINE_COL_CHECKLIST)
        if cl_raw and str(cl_raw).strip() not in ("", "nan", "None"):
            checklists = self._parse_checklist_json(str(cl_raw))

        deal = PipelineDeal(
            company_name      = company_name,
            instrument_type   = instrument,
            issuer_type       = issuer_type,
            asset_class       = asset_class,
            issuance_size     = issuance_size,
            funding_date      = funding_date,
            rating            = rating,
            security          = security,
            created_date      = created_date,
            status            = status,
            checklists        = checklists,
        )
        # if checklist is empty, initialise it
        if not deal.checklists:
            deal.checklists = self.initialize_checklist_for_deal(instrument)
        deal.update_checklist_progress()
        return deal

    # ══════════════════════════════════════════════
    #  READ CLOSED DEALS
    # ══════════════════════════════════════════════

    def load_closed_deals(self):
        try:
            df = pd.read_excel(self.file_path, sheet_name=config.SHEET_CLOSED)
        except Exception:
            return []

        if df.empty:
            return []

        deals = []
        for _, row in df.iterrows():
            try:
                deal = self._row_to_closed(row)
                if deal:
                    deals.append(deal)
            except Exception as e:
                print(f"[DataStore] skip closed row: {e}")
        return deals

    def _row_to_closed(self, row):
        import math

        def get(col, default=None):
            val = row.get(col)
            if isinstance(val, float) and math.isnan(val):
                return default
            return val if val is not None else default

        company_name = str(get(CLOSED_COL_ISSUER, "")).strip()
        if not company_name or company_name.lower() in ("nan", "none", ""):
            return None

        raw_type     = str(get(CLOSED_COL_TYPE, "Listed"))
        instrument   = normalise_instrument_type(raw_type)
        issuer_type  = str(get(CLOSED_COL_ISSUER_TYPE, "FS")).strip()
        asset_class  = "NBFC" if issuer_type == "FS" else "Corporate"

        raw_quantum  = get(CLOSED_COL_QUANTUM, 0)
        try:
            issuance_size = float(str(raw_quantum).replace(',', ''))
        except Exception:
            issuance_size = 0.0

        isin          = str(get(CLOSED_COL_ISIN, "N/A") or "N/A").strip()
        coupon        = parse_coupon(get(CLOSED_COL_COUPON, 0))
        tenor_raw     = get(CLOSED_COL_TENOR, 12)
        try:
            tenor = int(float(str(tenor_raw)))
        except Exception:
            tenor = 12
        rating        = str(get(CLOSED_COL_RATING, "N/A") or "N/A").strip()
        security      = str(get(CLOSED_COL_SECURITY, "N/A") or "N/A").strip()
        funding_date  = parse_excel_date(get(CLOSED_COL_ISSUE_DATE))
        maturity_date = parse_excel_date(get(CLOSED_COL_MATURITY))

        return ClosedDeal(
            company_name    = company_name,
            instrument_type = instrument,
            issuer_type     = issuer_type,
            asset_class     = asset_class,
            issuance_size   = issuance_size,
            isin            = isin,
            coupon          = coupon,
            tenor           = tenor,
            rating          = rating,
            security        = security,
            funding_date    = funding_date,
            maturity_date   = maturity_date,
        )

    # ══════════════════════════════════════════════
    #  WRITE PIPELINE
    # ══════════════════════════════════════════════

    def save_pipeline_deal(self, deal):
        wb    = load_workbook(self.file_path)
        sheet = self._get_or_create_sheet(wb, config.SHEET_PIPELINE, PIPELINE_HEADERS)
        row   = sheet.max_row + 1

        self._write_pipeline_row(sheet, row, deal)
        wb.save(self.file_path)

    def update_pipeline_deal(self, company_name, updated_deal):
        wb    = load_workbook(self.file_path)
        sheet = self._get_or_create_sheet(wb, config.SHEET_PIPELINE, PIPELINE_HEADERS)
        col_map = self._get_col_map(sheet)

        for r in range(2, sheet.max_row + 1):
            issuer_col = col_map.get(PIPELINE_COL_ISSUER, 1)
            if str(sheet.cell(r, issuer_col).value or "").strip().lower() == company_name.lower():
                self._write_pipeline_row(sheet, r, updated_deal, col_map)
                break

        wb.save(self.file_path)

    def delete_pipeline_deal(self, company_name):
        wb    = load_workbook(self.file_path)
        sheet = self._get_or_create_sheet(wb, config.SHEET_PIPELINE, PIPELINE_HEADERS)
        col_map = self._get_col_map(sheet)
        issuer_col = col_map.get(PIPELINE_COL_ISSUER, 1)

        for r in range(2, sheet.max_row + 1):
            if str(sheet.cell(r, issuer_col).value or "").strip().lower() == company_name.lower():
                sheet.delete_rows(r)
                break
        wb.save(self.file_path)

    def _write_pipeline_row(self, sheet, row_idx, deal, col_map=None):
        if col_map is None:
            col_map = self._get_col_map(sheet)

        cl_json = self._format_checklist_json(deal.checklists)

        values = {
            PIPELINE_COL_ISSUER:      deal.company_name,
            PIPELINE_COL_TYPE:        deal.instrument_type,
            PIPELINE_COL_ISSUER_TYPE: deal.issuer_type,
            PIPELINE_COL_QUANTUM:     deal.issuance_size,
            PIPELINE_COL_DATE:        deal.funding_date,
            PIPELINE_COL_CREDIT:      deal.status,
            PIPELINE_COL_RATING:      deal.rating,
            PIPELINE_COL_SECURITY:    deal.security,
            PIPELINE_COL_STATUS:      deal.status,
            PIPELINE_COL_CREATED:     deal.created_date,
            PIPELINE_COL_CHECKLIST:   cl_json,
        }

        for col_name, val in values.items():
            if col_name in col_map:
                sheet.cell(row_idx, col_map[col_name]).value = val
            else:
                # append new column
                next_col = sheet.max_column + 1
                sheet.cell(1, next_col).value = col_name
                sheet.cell(row_idx, next_col).value = val

    # ══════════════════════════════════════════════
    #  WRITE CLOSED
    # ══════════════════════════════════════════════

    def save_closed_deal(self, deal):
        wb    = load_workbook(self.file_path)
        sheet = self._get_or_create_sheet(wb, config.SHEET_CLOSED, CLOSED_HEADERS)
        row   = sheet.max_row + 1
        col_map = self._get_col_map(sheet)

        values = {
            CLOSED_COL_ISSUER:       deal.company_name,
            CLOSED_COL_TYPE:         deal.instrument_type,
            CLOSED_COL_ISSUER_TYPE:  deal.issuer_type,
            CLOSED_COL_QUANTUM:      deal.issuance_size,
            CLOSED_COL_ISIN:         deal.isin,
            CLOSED_COL_COUPON:       deal.coupon,
            CLOSED_COL_TENOR:        deal.tenor,
            CLOSED_COL_RATING:       deal.rating,
            CLOSED_COL_SECURITY:     deal.security,
            CLOSED_COL_ISSUE_DATE:   deal.funding_date,
            CLOSED_COL_MATURITY:     deal.maturity_date,
        }
        for col_name, val in values.items():
            if col_name in col_map:
                sheet.cell(row, col_map[col_name]).value = val
            else:
                next_col = sheet.max_column + 1
                sheet.cell(1, next_col).value = col_name
                sheet.cell(row, next_col).value = val

        wb.save(self.file_path)

    def move_to_closed(self, pipeline_deal, closed_deal):
        self.delete_pipeline_deal(pipeline_deal.company_name)
        self.save_closed_deal(closed_deal)

    # ══════════════════════════════════════════════
    #  LOOKUP HELPERS
    # ══════════════════════════════════════════════

    def company_exists(self, company_name):
        return any(d.company_name.lower() == company_name.lower()
                   for d in self.load_pipeline_deals())

    def get_deal_by_company(self, company_name):
        for d in self.load_pipeline_deals():
            if d.company_name.lower() == company_name.lower():
                return d
        return None

    # ══════════════════════════════════════════════
    #  CHECKLIST INIT
    # ══════════════════════════════════════════════

    def initialize_checklist_for_deal(self, instrument_type):
        items_map = get_checklist_for_instrument(instrument_type)
        checklists = {}
        for phase, items in items_map.items():
            pc = PhaseChecklist(phase_name=phase)
            for step_num, (task, maker, timing, listed_only) in enumerate(items, 1):
                pc.items.append(ChecklistItem(
                    step_number = step_num,
                    task_title  = task,
                    maker       = maker,
                    timing_note = timing,
                    listed_only = listed_only,
                ))
            checklists[phase] = pc
        return checklists

    # ══════════════════════════════════════════════
    #  INTERNAL HELPERS
    # ══════════════════════════════════════════════

    def _get_col_map(self, sheet):
        """Return {header_name: col_index} from row 1."""
        return {str(sheet.cell(1, c).value or "").strip(): c
                for c in range(1, sheet.max_column + 1)
                if sheet.cell(1, c).value}

    def _get_or_create_sheet(self, wb, sheet_name, headers):
        if sheet_name not in wb.sheetnames:
            sheet = wb.create_sheet(sheet_name)
            self._setup_sheet_headers(sheet, headers)
        return wb[sheet_name]

    def _format_checklist_json(self, checklists):
        return json.dumps({p: c.to_dict() for p, c in checklists.items()}, default=str)

    def _parse_checklist_json(self, json_str):
        try:
            data = json.loads(json_str)
            return {p: PhaseChecklist.from_dict(cd) for p, cd in data.items()}
        except Exception:
            return {}


# ══════════════════════════════════════════════════════════
#  GOOGLE SHEETS BACKEND
# ══════════════════════════════════════════════════════════

class GSheetDataStore(DataStore):
    """
    Same interface as DataStore but reads/writes Google Sheets.
    Inherits all parsing methods (_row_to_pipeline, _row_to_closed, etc.)
    and overrides only the I/O methods.

    Requirements: pip install gspread google-auth
    Credentials:  place service_account.json in Code_Streamlit/
    """

    def __init__(self):
        # DO NOT call super().__init__() — we skip file I/O setup
        self.file_path = config.DATA_FILE   # kept for compatibility; not used for I/O
        self._setup_gsheets()

    # ── setup ──────────────────────────────────────
    def _get_gspread_client(self):
        """
        Obtain an authorised gspread client.
        Priority:
          1. Local service_account.json file  (local dev or file uploaded to server)
          2. st.secrets['gcp_service_account'] (Streamlit Cloud Secrets)
        Raises FileNotFoundError if neither source is available.
        """
        import gspread

        # 1. Local file — works for local dev and for cloud if file is present
        if config.GOOGLE_CREDS_FILE.exists():
            return gspread.service_account(filename=str(config.GOOGLE_CREDS_FILE))

        # 2. Streamlit Secrets — for Streamlit Cloud deployment
        try:
            import streamlit as st
            if hasattr(st, 'secrets') and 'gcp_service_account' in st.secrets:
                creds_dict = {k: v for k, v in st.secrets['gcp_service_account'].items()}
                return gspread.service_account_from_dict(creds_dict)
        except Exception:
            pass

        raise FileNotFoundError(
            "No Google credentials found. Either:\n"
            "  • Place service_account.json in the Code_Streamlit/ folder, OR\n"
            "  • Add [gcp_service_account] to Streamlit Cloud → Settings → Secrets"
        )

    def _setup_gsheets(self):
        try:
            import gspread  # noqa: F401
        except ImportError:
            raise ImportError(
                "Google Sheets packages missing. Run:\n"
                "  pip install gspread google-auth"
            )

        self.gc = self._get_gspread_client()
        self.ss = self._open_or_create_spreadsheet()

    def _open_or_create_spreadsheet(self):
        """
        Open the Google Sheet by name.
        Auto-create is intentionally disabled: service accounts have no Drive
        storage quota, so gc.create() always raises a 403 quota error.
        Instead we show clear one-time setup instructions to the user.
        """
        import gspread
        try:
            return self.gc.open(config.GOOGLE_SHEET_NAME)
        except gspread.SpreadsheetNotFound:
            try:
                sa_email = self.gc.auth.service_account_email
            except Exception:
                sa_email = "your-service-account@project.iam.gserviceaccount.com"
            raise RuntimeError(
                f"Sheet '{config.GOOGLE_SHEET_NAME}' not found or not yet shared.\n\n"
                f"One-time setup (takes 1 minute):\n"
                f"1. Open drive.google.com (logged in as prabhu.ken01@gmail.com)\n"
                f"2. Click New → Google Sheets → rename it to exactly:  Issuance Tracker\n"
                f"3. Click Share → paste this email → give Editor access:\n"
                f"   {sa_email}\n"
                f"4. Refresh this app — columns & data will auto-populate."
            )

    def _ws(self, name):
        """Return worksheet by name; create it with headers if missing."""
        import gspread
        try:
            return self.ss.worksheet(name)
        except gspread.WorksheetNotFound:
            headers = PIPELINE_HEADERS if name == config.SHEET_PIPELINE else CLOSED_HEADERS
            ws = self.ss.add_worksheet(title=name, rows=500, cols=len(headers) + 5)
            ws.append_row(headers)
            return ws

    @property
    def spreadsheet_url(self):
        return f"https://docs.google.com/spreadsheets/d/{self.ss.id}"

    # ── READ ───────────────────────────────────────
    def load_pipeline_deals(self):
        ws = self._ws(config.SHEET_PIPELINE)
        records = ws.get_all_records(
            expected_headers=[],
            value_render_option='FORMATTED_VALUE'
        )
        deals = []
        for rec in records:
            try:
                row = pd.Series(rec)
                deal = self._row_to_pipeline(row)   # inherited from DataStore
                if deal:
                    deals.append(deal)
            except Exception as e:
                print(f"[GSheetDataStore] skip pipeline row: {e}")
        return deals

    def load_closed_deals(self):
        ws = self._ws(config.SHEET_CLOSED)
        records = ws.get_all_records(
            expected_headers=[],
            value_render_option='FORMATTED_VALUE'
        )
        deals = []
        for rec in records:
            try:
                row = pd.Series(rec)
                deal = self._row_to_closed(row)     # inherited from DataStore
                if deal:
                    deals.append(deal)
            except Exception as e:
                print(f"[GSheetDataStore] skip closed row: {e}")
        return deals

    # ── WRITE PIPELINE ─────────────────────────────
    def _pipeline_row_values(self, deal):
        """Return an ordered list of values matching the sheet's header row."""
        ws = self._ws(config.SHEET_PIPELINE)
        headers = ws.row_values(1)
        cl_json = self._format_checklist_json(deal.checklists)
        row_dict = {
            PIPELINE_COL_ISSUER:      deal.company_name,
            PIPELINE_COL_TYPE:        deal.instrument_type,
            PIPELINE_COL_ISSUER_TYPE: deal.issuer_type,
            PIPELINE_COL_QUANTUM:     deal.issuance_size,
            PIPELINE_COL_DATE:        str(deal.funding_date),
            PIPELINE_COL_CREDIT:      deal.status,
            PIPELINE_COL_RATING:      deal.rating,
            PIPELINE_COL_SECURITY:    deal.security,
            PIPELINE_COL_STATUS:      deal.status,
            PIPELINE_COL_CREATED:     str(deal.created_date),
            PIPELINE_COL_CHECKLIST:   cl_json,
        }
        return headers, [str(row_dict.get(h, '')) for h in headers]

    def save_pipeline_deal(self, deal):
        ws = self._ws(config.SHEET_PIPELINE)
        _, ordered = self._pipeline_row_values(deal)
        ws.append_row(ordered, value_input_option='USER_ENTERED')

    def update_pipeline_deal(self, company_name, updated_deal):
        ws = self._ws(config.SHEET_PIPELINE)
        headers, ordered = self._pipeline_row_values(updated_deal)
        try:
            issuer_col = headers.index(PIPELINE_COL_ISSUER) + 1
        except ValueError:
            return
        all_issuers = ws.col_values(issuer_col)
        for row_i, val in enumerate(all_issuers[1:], 2):   # skip header
            if str(val).strip().lower() == company_name.lower():
                ws.update(
                    range_name=f'A{row_i}',
                    values=[ordered],
                    value_input_option='USER_ENTERED'
                )
                return

    def delete_pipeline_deal(self, company_name):
        ws = self._ws(config.SHEET_PIPELINE)
        headers = ws.row_values(1)
        try:
            issuer_col = headers.index(PIPELINE_COL_ISSUER) + 1
        except ValueError:
            return
        all_issuers = ws.col_values(issuer_col)
        for row_i, val in enumerate(all_issuers[1:], 2):
            if str(val).strip().lower() == company_name.lower():
                ws.delete_rows(row_i)
                return

    # ── WRITE CLOSED ───────────────────────────────
    def save_closed_deal(self, deal):
        ws = self._ws(config.SHEET_CLOSED)
        headers = ws.row_values(1)
        row_dict = {
            CLOSED_COL_ISSUER:      deal.company_name,
            CLOSED_COL_TYPE:        deal.instrument_type,
            CLOSED_COL_ISSUER_TYPE: deal.issuer_type,
            CLOSED_COL_QUANTUM:     deal.issuance_size,
            CLOSED_COL_ISIN:        deal.isin,
            CLOSED_COL_COUPON:      deal.coupon,
            CLOSED_COL_TENOR:       deal.tenor,
            CLOSED_COL_RATING:      deal.rating,
            CLOSED_COL_SECURITY:    deal.security,
            CLOSED_COL_ISSUE_DATE:  str(deal.funding_date),
            CLOSED_COL_MATURITY:    str(deal.maturity_date),
        }
        ordered = [str(row_dict.get(h, '')) for h in headers]
        ws.append_row(ordered, value_input_option='USER_ENTERED')

    def move_to_closed(self, pipeline_deal, closed_deal):
        self.delete_pipeline_deal(pipeline_deal.company_name)
        self.save_closed_deal(closed_deal)

    # ── OVERRIDES that delegate to parent ──────────
    def company_exists(self, company_name):
        return any(d.company_name.lower() == company_name.lower()
                   for d in self.load_pipeline_deals())

    def get_deal_by_company(self, company_name):
        for d in self.load_pipeline_deals():
            if d.company_name.lower() == company_name.lower():
                return d
        return None


# ══════════════════════════════════════════════════════════
#  FACTORY
# ══════════════════════════════════════════════════════════

# Session-state key used to cache the GSheetDataStore instance.
# Shared between data_store.py and app.py so both use the same cache slot.
_GSHEET_CACHE_KEY = '_gsheet_store'


def get_data_store():
    """
    Return the right DataStore based on the storage_mode session flag.

    When Google Sheets mode is active the GSheetDataStore is cached in
    st.session_state[_GSHEET_CACHE_KEY] so that gc.open() (a counted Read
    API call) is only executed ONCE per browser session, not on every
    Streamlit rerun.  This prevents the 429 "Quota exceeded" errors that
    occur when the app rerenders rapidly (widget interactions, st.rerun()).

    Falls back to local Excel if Google Sheets fails or creds are missing.
    """
    try:
        import streamlit as st
        mode = st.session_state.get('storage_mode', '📁 Local Folder')
    except Exception:
        mode = '📁 Local Folder'

    if 'Google' in mode:
        if not config.google_creds_available():
            try:
                import streamlit as st
                st.warning(
                    "⚠️ Google Drive mode selected but no credentials found. "
                    "Using local Excel.\n\n"
                    "**To enable Google Sheets:** place `service_account.json` in the "
                    "Code_Streamlit/ folder, or add `[gcp_service_account]` to "
                    "Streamlit Cloud → Settings → Secrets."
                )
            except Exception:
                pass
            return DataStore()

        # ── Return cached instance if available ─────────────────────────────
        try:
            import streamlit as st
            cached = st.session_state.get(_GSHEET_CACHE_KEY)
            if cached is not None:
                return cached
        except Exception:
            pass

        # ── First call in this session: create and cache ─────────────────────
        try:
            store = GSheetDataStore()
            try:
                import streamlit as st
                st.session_state[_GSHEET_CACHE_KEY] = store
            except Exception:
                pass
            return store
        except RuntimeError as e:
            # RuntimeError = one-time setup needed (sheet not found / not shared)
            # Show as amber warning with actionable steps, not a red crash box
            try:
                import streamlit as st
                st.warning(f"☁️ Google Sheets — setup needed\n\n{e}\n\n"
                           f"Using local Excel in the meantime.")
            except Exception:
                print(f"[get_data_store] GSheets setup needed: {e}")
            return DataStore()
        except Exception as e:
            try:
                import streamlit as st
                err_str = str(e)
                if '429' in err_str:
                    st.warning(
                        "⚠️ Google Sheets rate limit hit (429 — too many read requests).  "
                        "Wait ~30 seconds then refresh the page."
                    )
                else:
                    st.error(f"Google Sheets connection failed: {e}\nFalling back to local Excel.")
                # Clear bad cache entry so next reload retries cleanly
                st.session_state.pop(_GSHEET_CACHE_KEY, None)
            except Exception:
                print(f"[get_data_store] Google Sheets failed: {e}. Using local Excel.")
            return DataStore()

    return DataStore()
