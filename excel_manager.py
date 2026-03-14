"""
NCD Tracker - Excel Manager
Handles all Excel file read/write operations
"""

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, datetime
from pathlib import Path
import json

import config
from field_definitions import *
from checklist_items import *
from data_models import PipelineDeal, ClosedDeal, PhaseChecklist, ChecklistItem
from helpers import parse_date, date_to_string

class ExcelManager:
    """Manages Excel file operations"""
    
    def __init__(self, file_path=None):
        """Initialize Excel Manager"""
        self.file_path = Path(file_path) if file_path else config.DATA_FILE
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        """Create Excel file if it doesn't exist"""
        if not self.file_path.exists():
            self._create_new_file()
    
    def _create_new_file(self):
        """Create new Excel file with proper structure"""
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Create Pipeline sheet
        pipeline_sheet = wb.create_sheet(config.SHEET_PIPELINE)
        self._setup_sheet_headers(pipeline_sheet, PIPELINE_HEADERS)
        
        # Create Closed Deals sheet
        closed_sheet = wb.create_sheet(config.SHEET_CLOSED)
        self._setup_sheet_headers(closed_sheet, CLOSED_HEADERS)
        
        wb.save(self.file_path)
    
    def _setup_sheet_headers(self, sheet, headers):
        """Setup headers with formatting"""
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = header
            
            # Apply header style
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Set column width
            width = COLUMN_WIDTHS.get(header, 15)
            sheet.column_dimensions[chr(64 + col_idx)].width = width
        
        # Freeze header row
        sheet.freeze_panes = 'A2'
    
    def load_pipeline_deals(self):
        """
        Load all pipeline deals from Excel
        
        Returns:
            list: List of PipelineDeal objects
        """
        try:
            df = pd.read_excel(self.file_path, sheet_name=config.SHEET_PIPELINE)
            
            if df.empty:
                return []
            
            deals = []
            for _, row in df.iterrows():
                try:
                    # Parse checklist progress if stored as JSON
                    checklists = {}
                    if 'Checklist Data' in row and pd.notna(row['Checklist Data']):
                        checklists = self._parse_checklist_data(row['Checklist Data'])
                    
                    deal = PipelineDeal(
                        company_name=str(row['Company Name']),
                        instrument_type=str(row['Instrument Type']),
                        asset_class=str(row['Asset Class']),
                        issuance_size=float(row['Issuance Size (₹ Cr)']),
                        funding_date=pd.to_datetime(row['Funding Date (T)']).date(),
                        rating=str(row['Rating']),
                        security=str(row['Security']),
                        checklist_progress=str(row['Checklist Progress']) if pd.notna(row.get('Checklist Progress')) else "0/0 - 0%",
                        created_date=pd.to_datetime(row['Created Date']).date() if pd.notna(row.get('Created Date')) else date.today(),
                        status=str(row['Status']) if pd.notna(row.get('Status')) else "In Progress",
                        checklists=checklists
                    )
                    deals.append(deal)
                except Exception as e:
                    print(f"Error loading deal row: {e}")
                    continue
            
            return deals
        
        except FileNotFoundError:
            return []
        except Exception as e:
            print(f"Error loading pipeline deals: {e}")
            return []
    
    def load_closed_deals(self):
        """
        Load all closed deals from Excel
        
        Returns:
            list: List of ClosedDeal objects
        """
        try:
            df = pd.read_excel(self.file_path, sheet_name=config.SHEET_CLOSED)
            
            if df.empty:
                return []
            
            deals = []
            for _, row in df.iterrows():
                try:
                    deal = ClosedDeal(
                        company_name=str(row['Company Name']),
                        instrument_type=str(row['Instrument Type']),
                        asset_class=str(row['Asset Class']),
                        issuance_size=float(row['Issuance Size (₹ Cr)']),
                        isin=str(row['ISIN']),
                        coupon=float(row['Coupon (% p.a.)']),
                        tenor=int(row['Tenor (Months)']),
                        rating=str(row['Rating']),
                        security=str(row['Security']),
                        funding_date=pd.to_datetime(row['Funding Date']).date(),
                        maturity_date=pd.to_datetime(row['Maturity Date']).date()
                    )
                    deals.append(deal)
                except Exception as e:
                    print(f"Error loading closed deal row: {e}")
                    continue
            
            return deals
        
        except FileNotFoundError:
            return []
        except Exception as e:
            print(f"Error loading closed deals: {e}")
            return []
    
    def save_pipeline_deal(self, deal):
        """
        Save a single pipeline deal to Excel
        
        Args:
            deal (PipelineDeal): Deal to save
        """
        wb = load_workbook(self.file_path)
        sheet = wb[config.SHEET_PIPELINE]
        
        # Find next empty row
        next_row = sheet.max_row + 1
        
        # Write data
        row_data = deal.to_excel_row()
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=next_row, column=col_idx)
            cell.value = value
            
            # Format based on data type
            self._format_cell(cell, PIPELINE_HEADERS[col_idx - 1])
        
        # Save checklist data as JSON comment or hidden column
        # For now, we'll use a hidden column approach
        
        wb.save(self.file_path)
    
    def save_closed_deal(self, deal):
        """
        Save a single closed deal to Excel
        
        Args:
            deal (ClosedDeal): Deal to save
        """
        wb = load_workbook(self.file_path)
        sheet = wb[config.SHEET_CLOSED]
        
        # Find next empty row
        next_row = sheet.max_row + 1
        
        # Write data
        row_data = deal.to_excel_row()
        for col_idx, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=next_row, column=col_idx)
            cell.value = value
            
            # Format based on data type
            self._format_cell(cell, CLOSED_HEADERS[col_idx - 1])
        
        wb.save(self.file_path)
    
    def update_pipeline_deal(self, company_name, updated_deal):
        """
        Update an existing pipeline deal
        
        Args:
            company_name (str): Company name to find
            updated_deal (PipelineDeal): Updated deal data
        """
        wb = load_workbook(self.file_path)
        sheet = wb[config.SHEET_PIPELINE]
        
        # Find row with matching company name
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == company_name:
                # Update row
                row_data = updated_deal.to_excel_row()
                for col_idx, value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    self._format_cell(cell, PIPELINE_HEADERS[col_idx - 1])
                break
        
        wb.save(self.file_path)
    
    def delete_pipeline_deal(self, company_name):
        """
        Delete a pipeline deal
        
        Args:
            company_name (str): Company name to delete
        """
        wb = load_workbook(self.file_path)
        sheet = wb[config.SHEET_PIPELINE]
        
        # Find and delete row
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == company_name:
                sheet.delete_rows(row_idx)
                break
        
        wb.save(self.file_path)
    
    def move_to_closed(self, pipeline_deal, closed_deal):
        """
        Move deal from pipeline to closed
        
        Args:
            pipeline_deal (PipelineDeal): Pipeline deal to remove
            closed_deal (ClosedDeal): Closed deal to add
        """
        # Delete from pipeline
        self.delete_pipeline_deal(pipeline_deal.company_name)
        
        # Add to closed
        self.save_closed_deal(closed_deal)
    
    def _format_cell(self, cell, header):
        """Apply formatting to cell based on header"""
        # Set alignment
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Set number format based on header
        if 'Size' in header or 'Amount' in header or '₹' in header:
            cell.number_format = NUMBER_FORMATS['Amount']
        elif 'Coupon' in header or '%' in header:
            cell.number_format = NUMBER_FORMATS['Percentage']
        elif 'Tenor' in header or 'Months' in header:
            cell.number_format = NUMBER_FORMATS['Integer']
        elif 'Date' in header:
            cell.number_format = NUMBER_FORMATS['Date']
    
    def _parse_checklist_data(self, json_str):
        """Parse checklist JSON data"""
        try:
            data = json.loads(json_str)
            checklists = {}
            for phase, checklist_data in data.items():
                checklists[phase] = PhaseChecklist.from_dict(checklist_data)
            return checklists
        except:
            return {}
    
    def initialize_checklist_for_deal(self, instrument_type):
        """
        Initialize empty checklist structure for a deal
        
        Args:
            instrument_type (str): "Listed NCD" or "Unlisted NCD"
        
        Returns:
            dict: Dictionary of PhaseChecklist objects
        """
        checklist_items = get_checklist_for_instrument(instrument_type)
        checklists = {}
        
        for phase, items in checklist_items.items():
            phase_checklist = PhaseChecklist(phase_name=phase)
            
            for step_num, (task, maker, timing, listed_only) in enumerate(items, start=1):
                item = ChecklistItem(
                    step_number=step_num,
                    task_title=task,
                    maker=maker,
                    timing_note=timing,
                    listed_only=listed_only
                )
                phase_checklist.items.append(item)
            
            checklists[phase] = phase_checklist
        
        return checklists
    
    def company_exists(self, company_name):
        """
        Check if company already exists in pipeline
        
        Args:
            company_name (str): Company name
        
        Returns:
            bool: True if exists
        """
        deals = self.load_pipeline_deals()
        return any(deal.company_name.lower() == company_name.lower() for deal in deals)
    
    def get_deal_by_company(self, company_name):
        """
        Get deal by company name
        
        Args:
            company_name (str): Company name
        
        Returns:
            PipelineDeal or None: Found deal or None
        """
        deals = self.load_pipeline_deals()
        for deal in deals:
            if deal.company_name.lower() == company_name.lower():
                return deal
        return None
